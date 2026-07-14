"""Integration tests for the Extractor orchestrator.

These tests drive the full pipeline end-to-end with real files on disk
(inside pytest's ``tmp_path`` sandbox) but stub the LLM so no network
call is made. The point is to check that PDFs are read, extractions
land in the right Excel cells, the checkpoint is written correctly,
and failures are isolated per document.
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import pytest
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from extractkit.config import Settings
from extractkit.exceptions import LLMError
from extractkit.extractor import Extractor, ExtractorConfig
from extractkit.schemas import (
    EXCEL_COLUMNS,
    StructuredFields,
    SynthesisFields,
)


def _write_pdf(path: Path, body: str) -> None:
    """Emit a tiny valid PDF with the given body text.

    Real articles are much larger, but for these tests any PDF that
    ``pypdf`` can open and extract text from is sufficient.
    """
    pdf = canvas.Canvas(str(path), pagesize=letter)
    pdf.drawString(72, 720, body)
    pdf.save()


def _write_template(path: Path) -> None:
    """Create a fresh Excel template with the canonical headers."""
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(list(EXCEL_COLUMNS))
    workbook.save(path)


@pytest.fixture
def pdf_folder(tmp_path: Path) -> Path:
    """A folder with three tiny PDFs, mimicking a small paper collection."""
    folder = tmp_path / "papers"
    folder.mkdir()
    for name in ("paper_a.pdf", "paper_b.pdf", "paper_c.pdf"):
        _write_pdf(folder / name, f"Body of {name}")
    return folder


@pytest.fixture
def template_path(tmp_path: Path) -> Path:
    """A ready-to-use Excel template for the run."""
    path = tmp_path / "template.xlsx"
    _write_template(path)
    return path


@pytest.fixture
def output_path(tmp_path: Path) -> Path:
    """Destination for the extractor's Excel output."""
    return tmp_path / "output.xlsx"


@pytest.fixture
def checkpoint_path(tmp_path: Path) -> Path:
    """Destination for the extractor's checkpoint file."""
    return tmp_path / "output.xlsx.checkpoint.json"


def _stub_llm(
    *,
    structured: StructuredFields | None = None,
    synthesis: SynthesisFields | None = None,
    structured_side_effect: list | None = None,
    synthesis_side_effect: list | None = None,
) -> MagicMock:
    """Build a fake LLMClient that returns pre-baked extraction results.

    Either ``structured`` / ``synthesis`` (single value returned every
    call) or ``*_side_effect`` (list of results, one per call) may be
    supplied; the latter lets tests exercise the failure-isolation path
    by raising an exception for a specific PDF.
    """
    stub = MagicMock()
    if structured_side_effect is not None:
        stub.extract_structured.side_effect = structured_side_effect
    else:
        stub.extract_structured.return_value = structured
    if synthesis_side_effect is not None:
        stub.extract_synthesis.side_effect = synthesis_side_effect
    else:
        stub.extract_synthesis.return_value = synthesis
    return stub


def test_run_succeeds_end_to_end(
    pdf_folder: Path,
    template_path: Path,
    output_path: Path,
    checkpoint_path: Path,
    structured_extraction_dict: dict[str, str],
    synthesis_extraction_dict: dict[str, str],
) -> None:
    """Three PDFs, three rows, checkpoint says all done."""
    settings = Settings(OPENAI_API_KEY="sk-test-not-real")
    llm = _stub_llm(
        structured=StructuredFields(**structured_extraction_dict),
        synthesis=SynthesisFields(**synthesis_extraction_dict),
    )
    config = ExtractorConfig(
        pdf_folder=pdf_folder,
        template_path=template_path,
        output_path=output_path,
        checkpoint_path=checkpoint_path,
    )

    result = Extractor(settings=settings, config=config, llm_client=llm).run()

    assert result.total_pdfs == 3
    assert result.succeeded_this_run == 3
    assert result.failed == 0
    assert output_path.exists()
    assert checkpoint_path.exists()


def test_resume_skips_completed_pdfs(
    pdf_folder: Path,
    template_path: Path,
    output_path: Path,
    checkpoint_path: Path,
    structured_extraction_dict: dict[str, str],
    synthesis_extraction_dict: dict[str, str],
) -> None:
    """A second run on the same output must skip already-processed PDFs."""
    settings = Settings(OPENAI_API_KEY="sk-test-not-real")
    llm = _stub_llm(
        structured=StructuredFields(**structured_extraction_dict),
        synthesis=SynthesisFields(**synthesis_extraction_dict),
    )
    config = ExtractorConfig(
        pdf_folder=pdf_folder,
        template_path=template_path,
        output_path=output_path,
        checkpoint_path=checkpoint_path,
    )

    # First run: everything succeeds.
    Extractor(settings=settings, config=config, llm_client=llm).run()

    # Second run: nothing new should happen.
    second = Extractor(settings=settings, config=config, llm_client=llm).run()
    assert second.total_pdfs == 3
    assert second.skipped_already_done == 3
    assert second.succeeded_this_run == 0
    assert second.failed == 0


def test_llm_failure_on_one_pdf_does_not_abort_run(
    pdf_folder: Path,
    template_path: Path,
    output_path: Path,
    checkpoint_path: Path,
    structured_extraction_dict: dict[str, str],
    synthesis_extraction_dict: dict[str, str],
) -> None:
    """A single-PDF failure is recorded but the run continues."""
    settings = Settings(OPENAI_API_KEY="sk-test-not-real")
    good_structured = StructuredFields(**structured_extraction_dict)
    good_synthesis = SynthesisFields(**synthesis_extraction_dict)

    # Sorted PDF order in ``pdf_folder`` is a, b, c; fail the second.
    llm = _stub_llm(
        structured_side_effect=[
            good_structured,
            LLMError("boom"),
            good_structured,
        ],
        synthesis_side_effect=[
            good_synthesis,
            good_synthesis,  # never reached because structured raised first
            good_synthesis,
        ],
    )
    config = ExtractorConfig(
        pdf_folder=pdf_folder,
        template_path=template_path,
        output_path=output_path,
        checkpoint_path=checkpoint_path,
    )

    result = Extractor(settings=settings, config=config, llm_client=llm).run()

    assert result.total_pdfs == 3
    assert result.succeeded_this_run == 2
    assert result.failed == 1
    assert len(result.failures) == 1


def test_empty_pdf_folder_returns_zero_totals(
    tmp_path: Path,
    template_path: Path,
    output_path: Path,
    checkpoint_path: Path,
) -> None:
    """Running against an empty folder should not raise; totals are zero."""
    empty_folder = tmp_path / "empty_papers"
    empty_folder.mkdir()

    settings = Settings(OPENAI_API_KEY="sk-test-not-real")
    llm = _stub_llm()
    config = ExtractorConfig(
        pdf_folder=empty_folder,
        template_path=template_path,
        output_path=output_path,
        checkpoint_path=checkpoint_path,
    )

    result = Extractor(settings=settings, config=config, llm_client=llm).run()
    assert result.total_pdfs == 0
    assert result.succeeded_this_run == 0
    assert result.failed == 0


def test_missing_template_raises(
    pdf_folder: Path,
    tmp_path: Path,
    output_path: Path,
    checkpoint_path: Path,
) -> None:
    """A missing template surfaces as a clear error before any API call."""
    settings = Settings(OPENAI_API_KEY="sk-test-not-real")
    llm = _stub_llm()
    config = ExtractorConfig(
        pdf_folder=pdf_folder,
        template_path=tmp_path / "does_not_exist.xlsx",
        output_path=output_path,
        checkpoint_path=checkpoint_path,
    )

    with pytest.raises(FileNotFoundError):
        Extractor(settings=settings, config=config, llm_client=llm).run()


def test_output_rows_are_aligned_to_excel_columns(
    pdf_folder: Path,
    template_path: Path,
    output_path: Path,
    checkpoint_path: Path,
    structured_extraction_dict: dict[str, str],
    synthesis_extraction_dict: dict[str, str],
) -> None:
    """Cells must end up under the matching header, not just in order."""
    settings = Settings(OPENAI_API_KEY="sk-test-not-real")
    llm = _stub_llm(
        structured=StructuredFields(**structured_extraction_dict),
        synthesis=SynthesisFields(**synthesis_extraction_dict),
    )
    config = ExtractorConfig(
        pdf_folder=pdf_folder,
        template_path=template_path,
        output_path=output_path,
        checkpoint_path=checkpoint_path,
    )
    Extractor(settings=settings, config=config, llm_client=llm).run()

    workbook = load_workbook(output_path)
    worksheet = workbook.active
    assert worksheet is not None
    rows = list(worksheet.iter_rows(values_only=True))
    header_row = list(rows[0])
    data_row = list(rows[1])

    title_index = header_row.index("Title")
    year_index = header_row.index("Year")

    assert data_row[title_index] == structured_extraction_dict["title"]
    assert data_row[year_index] == structured_extraction_dict["year"]