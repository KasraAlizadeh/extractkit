"""Tests for the Excel template reader and incremental writer."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from extractkit.excel_handler import append_row, count_data_rows, read_column_headers
from extractkit.exceptions import ExcelError
from extractkit.schemas import EXCEL_COLUMNS


def test_read_column_headers_matches_schema(template_xlsx: Path) -> None:
    headers = read_column_headers(template_xlsx)
    assert tuple(headers) == EXCEL_COLUMNS


def test_read_column_headers_missing_file_raises(tmp_path: Path) -> None:
    with pytest.raises(ExcelError):
        read_column_headers(tmp_path / "missing.xlsx")


def test_append_row_creates_output_from_template(
    template_xlsx: Path, sample_row: list[str], tmp_path: Path
) -> None:
    """First append should copy the template forward and add the row."""
    output = tmp_path / "output.xlsx"
    assert not output.exists()

    append_row(template_path=template_xlsx, output_path=output, row=sample_row)

    assert output.exists()
    workbook = load_workbook(output)
    worksheet = workbook.active
    assert worksheet is not None
    rows = list(worksheet.iter_rows(values_only=True))
    # First row is headers, second row is our appended data.
    assert tuple(rows[0]) == EXCEL_COLUMNS
    assert list(rows[1]) == sample_row


def test_append_row_grows_existing_output(
    template_xlsx: Path, sample_row: list[str], tmp_path: Path
) -> None:
    """Subsequent appends should add new rows without losing earlier ones."""
    output = tmp_path / "output.xlsx"
    append_row(template_path=template_xlsx, output_path=output, row=sample_row)
    second_row = [f"row2-{i}" for i in range(len(EXCEL_COLUMNS))]
    append_row(template_path=template_xlsx, output_path=output, row=second_row)

    workbook = load_workbook(output)
    worksheet = workbook.active
    assert worksheet is not None
    rows = list(worksheet.iter_rows(values_only=True))
    assert len(rows) == 3  # header + 2 data rows
    assert list(rows[1]) == sample_row
    assert list(rows[2]) == second_row


def test_count_data_rows_is_zero_when_output_absent(tmp_path: Path) -> None:
    assert count_data_rows(tmp_path / "nope.xlsx") == 0


def test_count_data_rows_excludes_header(
    template_xlsx: Path, sample_row: list[str], tmp_path: Path
) -> None:
    output = tmp_path / "output.xlsx"
    append_row(template_path=template_xlsx, output_path=output, row=sample_row)
    append_row(template_path=template_xlsx, output_path=output, row=sample_row)
    assert count_data_rows(output) == 2
