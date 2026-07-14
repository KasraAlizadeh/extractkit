"""Read the Excel template and write extracted rows atomically.

The template is a workbook whose first row lists the expected column
headers. On each successful extraction we open the current output (or
copy the template if the output does not yet exist), append one row,
save to a sibling temp file, and atomically replace the target. That
way a crash mid-write never leaves the user with a truncated file.
"""

from __future__ import annotations

import os
import re
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook

from extractkit.exceptions import ExcelError
from extractkit.schemas import EXCEL_COLUMNS


# openpyxl rejects any control character other than tab (\x09), line feed
# (\x0a), and carriage return (\x0d) — it raises IllegalCharacterError.
# LLM output sometimes contains stray control characters (vertical tabs,
# form feeds, escape sequences) that survived PDF text extraction; strip
# them before writing so a single bad character does not abort the run.
_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _sanitize_for_excel(value: str) -> str:
    """Remove characters openpyxl refuses to write to a cell.

    Also caps the length at Excel's per-cell limit (32,767 characters)
    so unusually long extractions don't fail at write time.

    Args:
        value: Raw cell value produced by extraction.

    Returns:
        A cleaned string safe for openpyxl.Worksheet.append.
    """
    if not isinstance(value, str):
        return value
    cleaned = _ILLEGAL_CHARS_RE.sub("", value)
    return cleaned[:32767]


def read_column_headers(template_path: Path) -> list[str]:
    """Return the header row of the template workbook.

    Args:
        template_path: Location of the template `.xlsx` file.

    Returns:
        List of header strings from row 1.

    Raises:
        ExcelError: If the file does not exist, cannot be opened, or
            has no header row.
    """
    if not template_path.exists():
        raise ExcelError(f"Template not found: {template_path}")

    try:
        workbook = load_workbook(template_path, read_only=True)
    except Exception as exc:
        raise ExcelError(f"Could not open template: {exc}") from exc

    try:
        worksheet = workbook.active
        if worksheet is None:
            raise ExcelError("Template has no active sheet")

        header_row = next(worksheet.iter_rows(values_only=True), None)
        if header_row is None:
            raise ExcelError("Template has no header row")

        return [str(cell) if cell is not None else "" for cell in header_row]
    finally:
        workbook.close()


def count_data_rows(output_path: Path) -> int:
    """Return the number of data rows (excluding the header) in a workbook.

    Zero if the file does not exist, is empty, or contains only the
    header row.

    Args:
        output_path: Location of the workbook.

    Returns:
        Count of rows below the header row.

    Raises:
        ExcelError: If the workbook exists but cannot be opened.
    """
    if not output_path.exists():
        return 0

    try:
        workbook = load_workbook(output_path, read_only=True)
    except Exception as exc:
        raise ExcelError(f"Could not open workbook: {exc}") from exc

    try:
        worksheet = workbook.active
        if worksheet is None:
            return 0
        total_rows = worksheet.max_row or 0
        return max(0, total_rows - 1)  # subtract the header row
    finally:
        workbook.close()


def validate_headers(template_path: Path) -> None:
    """Confirm the template headers match ``EXCEL_COLUMNS`` exactly.

    A mismatch means the template is out of date and would silently
    drop or misplace data — better to fail loud, up front.

    Args:
        template_path: Location of the template file to check.

    Raises:
        ExcelError: If the header row does not match ``EXCEL_COLUMNS``.
    """
    actual = read_column_headers(template_path)
    expected = list(EXCEL_COLUMNS)
    if actual != expected:
        missing = [h for h in expected if h not in actual]
        extra = [h for h in actual if h not in expected]
        details = []
        if missing:
            details.append(f"missing: {missing}")
        if extra:
            details.append(f"extra: {extra}")
        detail_str = "; ".join(details) if details else "column order differs"
        raise ExcelError(
            f"Template headers do not match expected columns ({detail_str}). "
            f"Regenerate the template from EXCEL_COLUMNS."
        )


def append_row(
    *,
    template_path: Path,
    output_path: Path,
    row: list[str],
) -> None:
    """Append one row to the output workbook, atomically.

    On first write, the output file does not yet exist, so we copy the
    template to seed it. On subsequent writes we open the existing
    output and append. Either way, the write goes to a sibling temp
    file and is then atomically renamed over the target so a crash
    mid-write never leaves the user with a half-written file.

    Args:
        template_path: Location of the template file (used to seed the
            output on the first write).
        output_path: Destination file for the appended row.
        row: Values to append, aligned to ``EXCEL_COLUMNS`` order.

    Raises:
        ExcelError: If reading or writing the workbook fails.
    """
    if not output_path.exists():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(template_path, output_path)

    try:
        workbook = load_workbook(output_path)
    except Exception as exc:
        raise ExcelError(f"Could not open output workbook: {exc}") from exc

    try:
        worksheet = workbook.active
        if worksheet is None:
            raise ExcelError("Workbook has no active sheet")

        # Sanitize each cell so a single control character in LLM output
        # cannot abort the whole write.
        worksheet.append([_sanitize_for_excel(cell) for cell in row])

        # Save to a sibling temp file, then atomically replace the target.
        # ``os.replace`` is atomic on POSIX and on Windows, so the output
        # file is never observed in a half-written state.
        temp_path = output_path.with_suffix(output_path.suffix + ".tmp")
        try:
            workbook.save(temp_path)
        except Exception as exc:
            raise ExcelError(f"Could not save output workbook: {exc}") from exc

        os.replace(temp_path, output_path)
    finally:
        workbook.close()


def ensure_output_seeded(template_path: Path, output_path: Path) -> None:
    """Copy the template to the output path if the output is missing.

    Useful when the caller wants the output file to exist before any
    extractions succeed, so downstream tooling can open it early.

    Args:
        template_path: Location of the template file.
        output_path: Destination path to seed.
    """
    if not output_path.exists():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(template_path, output_path)


def create_blank_workbook_from_columns(
    output_path: Path,
    columns: tuple[str, ...] = EXCEL_COLUMNS,
) -> None:
    """Create a fresh workbook containing only the header row.

    Not used by the extraction pipeline itself, but handy for scripts
    that need to regenerate the template.

    Args:
        output_path: Where to save the new workbook.
        columns: Column headers to write to row 1.
    """
    workbook = Workbook()
    worksheet = workbook.active
    if worksheet is None:
        raise ExcelError("Newly created workbook has no active sheet")
    worksheet.title = "Extractions"
    worksheet.append(list(columns))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)